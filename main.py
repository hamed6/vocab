from datetime import datetime
import  openpyxl as opxl

# print(datetime.date.today())
# f=open('Book.xlsx','r')
# print(f.read())
# f.close()


wb=opxl.load_workbook('Book.xlsx')
ws=wb['Sheet1']
# c=ws['A1':'A4']
# print(wb['A1'.value])
# read
counter = 0
colNumber=ws.max_column

# while counter<10:        
#     for row in ws.iter_cols( min_col=colNumber, max_col=colNumber):
#         for cell in row:
#             if ( cell.value != None):
#                 word='A'+str(cell.row)  
               
#                 print(counter, ': ', ws[word].value)
#                 counter+=1     
#     colNumber-=1
#     if (colNumber==1):
#         break
    
# //////////////////


for row in ws.iter_rows():
    for r in row:
        if (r.value != None) :   # (r.value == datetime.date.today()):
            try:
                # dateObj = datetime.strptime( r.value,'%d/%m/%y')
                # print ( dateObj.date())
                print (type(r.value))
            except ValueError:
                print ('error')
                continue
